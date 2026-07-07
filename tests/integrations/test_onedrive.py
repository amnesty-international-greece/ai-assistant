"""Unit tests for OneDriveClient.

Mocks ``httpx`` and ``msal`` so the tests run hermetically with no network or
authentication side-effects.  Covers:
  - Path helpers (URL-encoding, archive_root joining)
  - Token acquisition + the OneDriveAuthRequired contract
  - Site-id caching (one network call, not many)
  - Public CRUD: upload_file, list_files, download_file, get_share_link
  - Protocol-registry helpers: get_next_protocol_number, append_protocol_row

Integration tests against the live SharePoint site live in
``scripts/`` and are run manually - anything that hits the network
stays out of CI.
"""
from __future__ import annotations

import ast
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import httpx
import pytest

from src.integrations.onedrive import OneDriveAuthRequired, OneDriveClient


# ── Fixtures ──────────────────────────────────────────────────────────────────


@pytest.fixture
def patched_settings():
    """Patch the module-level settings reference used inside onedrive.py."""
    # Patches against the NEW module path after the m365/ consolidation.
    # ``src.integrations.onedrive`` is now a re-export shim - patching there
    # mutates only the shim's namespace and the actual code in m365/onedrive.py
    # would silently bypass the patch.
    with patch("src.integrations.m365.onedrive.settings") as s:
        s.ms_client_id = "test-client"
        s.ms_client_secret = "test-secret"
        s.ms_tenant_id = "test-tenant"
        s.onedrive.sharepoint_host = "test.sharepoint.com"
        s.onedrive.sharepoint_site_path = "/sites/Test"
        s.onedrive.archive_root = "Αρχείο"
        s.onedrive.yearly_subfolder = "Αρχείο ανά έτος"
        s.onedrive.protocol_excel = "[Πρωτόκολλο] Αρχείο ΔΣ.xlsx"
        yield s


@pytest.fixture
def patched_msal():
    """Patch MSAL so OneDriveClient can be instantiated without a real cache."""
    # MSAL is now consumed inside the shared M365GraphAuthMixin in m365/auth.py.
    with patch("src.integrations.m365.auth.msal") as m:
        cache = MagicMock()
        cache.has_state_changed = False
        m.SerializableTokenCache.return_value = cache
        app = MagicMock()
        m.ConfidentialClientApplication.return_value = app
        yield m, app


@pytest.fixture
def patched_tokens():
    with (
        # token cache helpers live in m365/auth.py (the mixin module) now
        patch("src.integrations.m365.auth.get_section", return_value={}),
        patch("src.integrations.m365.auth.set_section") as set_section,
    ):
        yield set_section


@pytest.fixture(autouse=True)
def _isolated_protocol_backup(tmp_path, monkeypatch):
    """Redirect PROTOCOL_BACKUP_PATH to a per-test temp dir.

    Without this, the shared backup file at data/backups/protokollo_latest.xlsx
    from a prior real run would be returned by _workbook_path_for_read,
    contaminating every read test with stale rows.
    """
    monkeypatch.setattr(
        OneDriveClient,
        "PROTOCOL_BACKUP_PATH",
        tmp_path / "backups" / "protokollo_latest.xlsx",
    )
    yield


@pytest.fixture
def client(patched_settings, patched_msal, patched_tokens) -> OneDriveClient:
    _, app = patched_msal
    # Default: a valid cached account with a token
    app.get_accounts.return_value = [{"username": "test@example.com"}]
    app.acquire_token_silent.return_value = {"access_token": "test-token"}
    return OneDriveClient()


# ── Module / static helpers ───────────────────────────────────────────────────


def test_onedrive_module_parses() -> None:
    """Compile check (catches syntax errors before pytest collects)."""
    # parents[2] = project root (test file is at tests/integrations/foo.py)
    src = Path(__file__).resolve().parents[2] / "src" / "integrations" / "onedrive.py"
    ast.parse(src.read_text(encoding="utf-8"))


def test_encode_path_preserves_forward_slashes(client: OneDriveClient) -> None:
    """Greek characters are percent-encoded; forward slashes are not."""
    encoded = client._encode_path("Αρχείο/Αρχείο ανά έτος/2026")
    # Greek 'Α' → %CE%91, slashes preserved
    assert "/" in encoded
    assert encoded.startswith("%CE%91")
    assert "%CE%91%CF%81%CF%87%CE%B5%CE%AF%CE%BF/" in encoded


def test_archive_path_joins_components(client: OneDriveClient) -> None:
    """archive_root is prepended; empty parts are dropped."""
    assert client._archive_path("foo", "bar.pdf") == "Αρχείο/foo/bar.pdf"
    assert client._archive_path("", "bar.pdf") == "Αρχείο/bar.pdf"
    assert client._archive_path("//foo//", "bar.pdf") == "Αρχείο/foo/bar.pdf"


# ── Token contract ────────────────────────────────────────────────────────────


def test_get_token_raises_when_no_account(
    patched_settings, patched_msal, patched_tokens
) -> None:
    _, app = patched_msal
    app.get_accounts.return_value = []  # no cached account
    c = OneDriveClient()
    with pytest.raises(OneDriveAuthRequired, match="Run: ai-assistant auth microsoft"):
        c._get_token()


def test_get_token_raises_when_silent_refresh_fails(
    patched_settings, patched_msal, patched_tokens
) -> None:
    _, app = patched_msal
    app.get_accounts.return_value = [{"username": "x"}]
    app.acquire_token_silent.return_value = None  # MSAL signals refresh failure
    c = OneDriveClient()
    with pytest.raises(OneDriveAuthRequired, match="refresh failed"):
        c._get_token()


def test_get_token_returns_token_and_persists_cache(
    patched_settings, patched_msal, patched_tokens
) -> None:
    _, app = patched_msal
    app.get_accounts.return_value = [{"username": "x"}]
    app.acquire_token_silent.return_value = {"access_token": "abc-123"}

    # Mark cache as changed so _persist_cache fires
    cache = patched_msal[0].SerializableTokenCache.return_value
    cache.has_state_changed = True
    cache.serialize.return_value = '{"k":"v"}'

    c = OneDriveClient()
    token = c._get_token()
    assert token == "abc-123"
    patched_tokens.assert_called_once_with("microsoft", {"k": "v"})


# ── Site-id caching ───────────────────────────────────────────────────────────


def test_site_id_is_resolved_and_cached(client: OneDriveClient) -> None:
    """First call hits Graph; subsequent calls use the in-memory cache."""
    mock_resp = MagicMock()
    mock_resp.json.return_value = {"id": "site-xyz"}
    mock_resp.raise_for_status = MagicMock()

    with patch("src.integrations.m365.onedrive.httpx.Client") as mock_client_cls:
        mock_client = MagicMock()
        mock_client.get.return_value = mock_resp
        mock_client_cls.return_value.__enter__.return_value = mock_client

        assert client._get_site_id() == "site-xyz"
        assert client._get_site_id() == "site-xyz"  # second call

    # Only one network call despite two invocations
    assert mock_client.get.call_count == 1
    url = mock_client.get.call_args.args[0]
    assert "/sites/test.sharepoint.com:/sites/Test" in url


# ── Public CRUD ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_upload_file_targets_correct_sharepoint_path(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """upload_file PUTs to /sites/{id}/drive/root:/{archive_root}/{folder}/{filename}:/content."""
    local = tmp_path / "hello.pdf"
    local.write_bytes(b"%PDF-1.4 content")

    client._site_id = "site-xyz"  # short-circuit site resolution

    mock_resp = MagicMock()
    mock_resp.json.return_value = {"id": "item-1", "size": 16}
    mock_resp.raise_for_status = MagicMock()

    fake_async_client = MagicMock()
    fake_async_client.put = AsyncMock(return_value=mock_resp)

    with patch("src.integrations.m365.onedrive.httpx.AsyncClient") as ac:
        ac.return_value.__aenter__.return_value = fake_async_client

        result = await client.upload_file(
            local_path=local,
            remote_folder="Αρχείο ανά έτος/2026",
            filename="test.pdf",
        )

    assert result == {"id": "item-1", "size": 16}
    url = fake_async_client.put.call_args.kwargs.get("url") or fake_async_client.put.call_args.args[0]
    # URL should contain the encoded archive path and end with :/content
    assert "/sites/site-xyz/drive/root:/" in url
    assert url.endswith(":/content")
    # Greek characters should be percent-encoded
    assert "%CE%91%CF%81%CF%87%CE%B5%CE%AF%CE%BF" in url


@pytest.mark.asyncio
async def test_list_files_empty_path_lists_archive_root(client: OneDriveClient) -> None:
    """list_files('') queries the archive root, not an empty path."""
    client._site_id = "site-xyz"

    mock_resp = MagicMock()
    mock_resp.json.return_value = {"value": [{"name": "f1"}]}
    mock_resp.raise_for_status = MagicMock()

    fake_async_client = MagicMock()
    fake_async_client.get = AsyncMock(return_value=mock_resp)

    with patch("src.integrations.m365.onedrive.httpx.AsyncClient") as ac:
        ac.return_value.__aenter__.return_value = fake_async_client

        items = await client.list_files("")

    assert items == [{"name": "f1"}]
    url = fake_async_client.get.call_args.args[0]
    # archive_root in the URL, plus :/children
    assert "%CE%91%CF%81%CF%87%CE%B5%CE%AF%CE%BF" in url
    assert ":/children" in url


@pytest.mark.asyncio
async def test_get_share_link_returns_web_url(client: OneDriveClient) -> None:
    client._site_id = "site-xyz"

    mock_resp = MagicMock()
    mock_resp.json.return_value = {"link": {"webUrl": "https://share.example/x"}}
    mock_resp.raise_for_status = MagicMock()

    fake_async_client = MagicMock()
    fake_async_client.post = AsyncMock(return_value=mock_resp)

    with patch("src.integrations.m365.onedrive.httpx.AsyncClient") as ac:
        ac.return_value.__aenter__.return_value = fake_async_client

        url = await client.get_share_link("item-1")

    assert url == "https://share.example/x"
    posted_json = fake_async_client.post.call_args.kwargs["json"]
    assert posted_json == {"type": "view", "scope": "organization"}


# ── Delete (file + folder) ───────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_delete_file_targets_correct_sharepoint_path(client: OneDriveClient) -> None:
    """delete_file DELETEs /sites/{id}/drive/root:/{archive_root}/{path}."""
    client._site_id = "site-xyz"

    mock_resp = MagicMock()
    mock_resp.status_code = 204
    mock_resp.raise_for_status = MagicMock()

    fake_async_client = MagicMock()
    fake_async_client.delete = AsyncMock(return_value=mock_resp)

    with patch("src.integrations.m365.onedrive.httpx.AsyncClient") as ac:
        ac.return_value.__aenter__.return_value = fake_async_client
        await client.delete_file("Αρχείο ανά έτος/2026/foo.pdf")

    url = fake_async_client.delete.call_args.args[0]
    assert "/sites/site-xyz/drive/root:/" in url
    assert "%CE%91%CF%81%CF%87%CE%B5%CE%AF%CE%BF" in url   # Αρχείο encoded
    # No :/content or :/children suffix - the item URL itself
    assert not url.endswith(":/content")
    assert not url.endswith(":/children")


@pytest.mark.asyncio
async def test_delete_folder_uses_same_endpoint(client: OneDriveClient) -> None:
    """delete_folder hits the same DELETE endpoint as delete_file."""
    client._site_id = "site-xyz"

    mock_resp = MagicMock()
    mock_resp.status_code = 204
    mock_resp.raise_for_status = MagicMock()

    fake_async_client = MagicMock()
    fake_async_client.delete = AsyncMock(return_value=mock_resp)

    with patch("src.integrations.m365.onedrive.httpx.AsyncClient") as ac:
        ac.return_value.__aenter__.return_value = fake_async_client
        await client.delete_folder("_e2e_test")

    url = fake_async_client.delete.call_args.args[0]
    assert "/sites/site-xyz/drive/root:/" in url
    assert "_e2e_test" in url


# ── Protocol-registry helpers (Excel) ────────────────────────────────────────


@pytest.mark.asyncio
async def test_get_next_protocol_number_increments_last_entry(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """Reads the year sheet's last YYYY_NNN value and increments by one."""
    import openpyxl

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("2026")
    ws.append(["Πρωτόκολλο", "Ημερομηνία", "Έγγραφο"])
    ws.append(["2026_024", None, "x"])
    ws.append(["2026_025", None, "y"])
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    with patch.object(client, "download_file", side_effect=fake_download):
        next_id = await client.get_next_protocol_number(2026)

    assert next_id == "2026_026"


@pytest.mark.asyncio
async def test_get_next_protocol_number_year_rollover(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """If the year sheet is missing, start the sequence at 001."""
    import openpyxl

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("2025")
    ws.append(["Πρωτόκολλο"])
    ws.append(["2025_088"])
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    with patch.object(client, "download_file", side_effect=fake_download):
        next_id = await client.get_next_protocol_number(2027)

    assert next_id == "2027_001"


@pytest.mark.asyncio
async def test_get_next_protocol_number_empty_sheet(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """Year sheet exists but has only a header → start at 001."""
    import openpyxl

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("2026")
    ws.append(["Πρωτόκολλο", "Ημερομηνία"])
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    with patch.object(client, "download_file", side_effect=fake_download):
        next_id = await client.get_next_protocol_number(2026)

    assert next_id == "2026_001"


@pytest.mark.asyncio
async def test_append_protocol_row_writes_and_reuploads(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """append_protocol_row downloads → mutates → re-uploads under the same name."""
    import openpyxl

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("2026")
    ws.append(["Πρωτόκολλο", "Ημερομηνία", "Έγγραφο", "Κύρια Σημεία", "Ετικέτες"])
    ws.append(["2026_001", None, "old", None, "Διοικητικά"])
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    uploaded: dict = {}

    async def fake_upload(
        local_path: Path, remote_folder: str, filename: str | None = None, workflow: str = "x"
    ) -> dict:
        uploaded["folder"] = remote_folder
        uploaded["filename"] = filename
        uploaded["bytes"] = local_path.read_bytes()
        return {"id": "item-1"}

    with (
        patch.object(client, "download_file", side_effect=fake_download),
        patch.object(client, "upload_file", side_effect=fake_upload),
    ):
        await client.append_protocol_row(
            protocol_id="2026_002",
            date_str="2026-03-15",
            title="Πρακτικά - Συνεδρίαση ΔΣ02-2026",
            main_points="1. Foo\n2. Bar",
            tags="Διοικητικά, Πρακτικά",
        )

    # Re-uploaded to the archive root with the configured filename
    assert uploaded["folder"] == ""
    assert uploaded["filename"] == "[Πρωτόκολλο] Αρχείο ΔΣ.xlsx"

    # Open the bytes we re-uploaded and check the new row landed
    written = tmp_path / "written.xlsx"
    written.write_bytes(uploaded["bytes"])
    wb2 = openpyxl.load_workbook(written, data_only=True)
    ws2 = wb2["2026"]
    rows = list(ws2.iter_rows(values_only=True))
    # 1 header + 1 existing + 1 new
    assert len(rows) == 3
    new_row = rows[-1]
    assert new_row[0] == "2026_002"
    assert new_row[2] == "Πρακτικά - Συνεδρίαση ΔΣ02-2026"
    assert new_row[3] == "1. Foo\n2. Bar"
    assert new_row[4] == "Διοικητικά, Πρακτικά"


@pytest.mark.asyncio
async def test_append_protocol_row_creates_year_sheet_if_missing(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """Appending into a year that has no sheet yet creates it with headers."""
    import openpyxl

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("2025")  # only 2025 exists
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    uploaded_bytes: bytes = b""

    async def fake_upload(
        local_path: Path, remote_folder: str, filename: str | None = None, workflow: str = "x"
    ) -> dict:
        nonlocal uploaded_bytes
        uploaded_bytes = local_path.read_bytes()
        return {"id": "item-1"}

    with (
        patch.object(client, "download_file", side_effect=fake_download),
        patch.object(client, "upload_file", side_effect=fake_upload),
    ):
        await client.append_protocol_row(
            protocol_id="2026_001",
            date_str="2026-01-15",
            title="First of the year",
            tags="Διοικητικά",
        )

    written = tmp_path / "written.xlsx"
    written.write_bytes(uploaded_bytes)
    wb2 = openpyxl.load_workbook(written, data_only=True)
    assert "2026" in wb2.sheetnames
    ws2 = wb2["2026"]
    rows = list(ws2.iter_rows(values_only=True))
    # Header + one row
    assert rows[0] == ("Πρωτόκολλο", "Ημερομηνία", "Έγγραφο", "Κύρια Σημεία", "Ετικέτες")
    assert rows[1][0] == "2026_001"
    assert rows[1][2] == "First of the year"


@pytest.mark.asyncio
async def test_append_protocol_row_parses_date_string(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """date_str='YYYY-MM-DD' lands as a real date cell, not a string."""
    import openpyxl
    from datetime import date

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("2026")
    ws.append(["Πρωτόκολλο", "Ημερομηνία", "Έγγραφο"])
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    uploaded_bytes: bytes = b""

    async def fake_upload(
        local_path: Path, remote_folder: str, filename: str | None = None, workflow: str = "x"
    ) -> dict:
        nonlocal uploaded_bytes
        uploaded_bytes = local_path.read_bytes()
        return {"id": "item-1"}

    with (
        patch.object(client, "download_file", side_effect=fake_download),
        patch.object(client, "upload_file", side_effect=fake_upload),
    ):
        await client.append_protocol_row(
            protocol_id="2026_010",
            date_str="2026-05-23",
            title="t",
        )

    written = tmp_path / "written.xlsx"
    written.write_bytes(uploaded_bytes)
    wb2 = openpyxl.load_workbook(written)
    cell = wb2["2026"].cell(row=2, column=2)
    # openpyxl round-trips dates as datetime at midnight - compare on .date()
    assert hasattr(cell.value, "year") and cell.value.year == 2026
    assert cell.value.month == 5 and cell.value.day == 23


# ── tokens module ─────────────────────────────────────────────────────────────


def test_tokens_module_parses() -> None:
    """Compile check for the unified token store."""
    src = Path(__file__).resolve().parents[2] / "src" / "core" / "tokens.py"
    ast.parse(src.read_text(encoding="utf-8"))


# ── delete_protocol_row ──────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_delete_protocol_row_removes_matching_row(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """The matching row is removed and the workbook re-uploaded."""
    import openpyxl

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("2026")
    ws.append(["Πρωτόκολλο", "Ημερομηνία", "Έγγραφο", "Κύρια Σημεία", "Ετικέτες"])
    ws.append(["2026_001", None, "keep", None, None])
    ws.append(["2026_002", None, "remove me", None, None])
    ws.append(["2026_003", None, "keep", None, None])
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    uploaded_bytes: bytes = b""

    async def fake_upload(
        local_path: Path, remote_folder: str, filename: str | None = None, workflow: str = "x"
    ) -> dict:
        nonlocal uploaded_bytes
        uploaded_bytes = local_path.read_bytes()
        return {"id": "item-1"}

    with (
        patch.object(client, "download_file", side_effect=fake_download),
        patch.object(client, "upload_file", side_effect=fake_upload),
    ):
        result = await client.delete_protocol_row("2026_002")

    assert result is True

    # The remaining sheet should have header + 2 rows (001, 003)
    written = tmp_path / "written.xlsx"
    written.write_bytes(uploaded_bytes)
    wb2 = openpyxl.load_workbook(written, data_only=True)
    rows = list(wb2["2026"].iter_rows(values_only=True))
    assert len(rows) == 3
    protocol_ids = [r[0] for r in rows[1:]]
    assert "2026_002" not in protocol_ids
    assert protocol_ids == ["2026_001", "2026_003"]


@pytest.mark.asyncio
async def test_delete_protocol_row_returns_false_for_missing_id(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """Returns False (no-op) when the protocol_id doesn't exist in the sheet."""
    import openpyxl

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("2026")
    ws.append(["Πρωτόκολλο"])
    ws.append(["2026_001"])
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    with patch.object(client, "download_file", side_effect=fake_download):
        result = await client.delete_protocol_row("2026_999")

    assert result is False


# ── Taxonomy / categories / recent entries (Phase 1 archive) ────────────────


def _build_taxonomy_xlsx(path: Path) -> None:
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_tags = wb.create_sheet("Ετικέτες")
    ws_tags.append(["Ετικέτα", "Περιγραφή"])
    ws_tags.append(["Διοικητικά", "Διοικητικά έγγραφα του ΔΣ"])
    ws_tags.append(["Πρακτικά", "Πρακτικά συνεδριάσεων"])
    ws_tags.append(["Επιχειρησιακά", "Χρήση με φειδώ"])

    ws_cats = wb.create_sheet("Κατηγορίες")
    ws_cats.append(["Πρότυπο τίτλου", "Προεπιλεγμένες Ετικέτες", "Κύρια Σημεία"])
    ws_cats.append(["Πρακτικά - Συνεδρίαση ΔΣ##-####", "Διοικητικά, Πρακτικά", "Αριθμημένη λίστα"])
    ws_cats.append(["Πρόσκληση - Συνεδρίαση ΔΣ##-####", "Διοικητικά, Προσκλήσεις", "Ημερήσια διάταξη"])

    ws_2026 = wb.create_sheet("2026")
    ws_2026.append(["Πρωτόκολλο", "Ημερομηνία", "Έγγραφο", "Κύρια Σημεία", "Ετικέτες"])
    ws_2026.append(["2026_001", None, "Πρακτικά 1", "kp1", "Διοικητικά"])
    ws_2026.append(["2026_002", None, "Πρακτικά 2", "kp2", "Πρακτικά"])
    wb.save(path)


@pytest.mark.asyncio
async def test_read_taxonomy_skips_header(client: OneDriveClient, tmp_path: Path) -> None:
    fake_xlsx = tmp_path / "p.xlsx"
    _build_taxonomy_xlsx(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    with patch.object(client, "download_file", side_effect=fake_download):
        rows = await client.read_taxonomy()

    assert {"tag": "Διοικητικά", "description": "Διοικητικά έγγραφα του ΔΣ"} in rows
    assert any(r["tag"] == "Επιχειρησιακά" for r in rows)
    assert all(r["tag"] != "Ετικέτα" for r in rows)  # header skipped
    assert len(rows) == 3


@pytest.mark.asyncio
async def test_read_categories_skips_header(client: OneDriveClient, tmp_path: Path) -> None:
    fake_xlsx = tmp_path / "p.xlsx"
    _build_taxonomy_xlsx(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    with patch.object(client, "download_file", side_effect=fake_download):
        rows = await client.read_categories()

    assert len(rows) == 2
    assert rows[0]["pattern"].startswith("Πρακτικά")
    assert rows[0]["tags"] == "Διοικητικά, Πρακτικά"
    assert rows[0]["kuria_simeia"] == "Αριθμημένη λίστα"


@pytest.mark.asyncio
async def test_read_recent_entries_returns_last_n(client: OneDriveClient, tmp_path: Path) -> None:
    fake_xlsx = tmp_path / "p.xlsx"
    _build_taxonomy_xlsx(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    # Patch the date import used inside read_recent_entries so the test is
    # year-stable.  The method does `from datetime import date as _date` at
    # call time, so we patch the datetime module's `date` attribute.
    from datetime import date as _date
    import datetime as _datetime_module

    class _FrozenDate(_date):
        @classmethod
        def today(cls) -> "_date":
            return _date(2026, 6, 1)

    with patch.object(client, "download_file", side_effect=fake_download), \
         patch.object(_datetime_module, "date", _FrozenDate):
        entries = await client.read_recent_entries(n=5, years_back=1)

    assert len(entries) <= 5
    assert any(e["proto"] == "2026_001" for e in entries)
    assert any(e["proto"] == "2026_002" for e in entries)


@pytest.mark.asyncio
async def test_get_current_year_max_seq(client: OneDriveClient, tmp_path: Path) -> None:
    fake_xlsx = tmp_path / "p.xlsx"
    _build_taxonomy_xlsx(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    with patch.object(client, "download_file", side_effect=fake_download):
        max_seq = await client.get_current_year_max_seq(2026)
        zero = await client.get_current_year_max_seq(2099)

    assert max_seq == 2
    assert zero == 0


@pytest.mark.asyncio
async def test_delete_protocol_row_returns_false_for_missing_year_sheet(
    client: OneDriveClient, tmp_path: Path
) -> None:
    """Returns False when the year sheet doesn't exist at all."""
    import openpyxl

    fake_xlsx = tmp_path / "protokollo.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("2025")
    wb.save(fake_xlsx)

    async def fake_download(remote_path: str, local_path: Path) -> Path:
        local_path.write_bytes(fake_xlsx.read_bytes())
        return local_path

    with patch.object(client, "download_file", side_effect=fake_download):
        result = await client.delete_protocol_row("2026_001")

    assert result is False


# ── 423 Locked retry contract (regression for 2026-05-27 incident) ────────


@pytest.mark.asyncio
async def test_upload_protocol_workbook_retries_on_423_then_succeeds(
    client: OneDriveClient, tmp_path: Path, monkeypatch,
) -> None:
    """If SharePoint returns 423 Locked once, the helper waits + retries."""
    from src.integrations.m365.onedrive import _PROTOKOLLO_LOCK_RETRY_DELAYS

    # Don't actually sleep during the test
    async def _no_sleep(_): return None
    monkeypatch.setattr("asyncio.sleep", _no_sleep)

    call_count = 0

    async def fake_upload(*, local_path, remote_folder, filename, workflow):
        nonlocal call_count
        call_count += 1
        if call_count == 1:
            resp = httpx.Response(status_code=423, request=httpx.Request("PUT", "https://x"))
            raise httpx.HTTPStatusError("423", request=resp.request, response=resp)
        return {"id": "ok"}

    fake_xlsx = tmp_path / "p.xlsx"
    fake_xlsx.write_bytes(b"x")

    with patch.object(client, "upload_file", side_effect=fake_upload):
        result = await client._upload_protocol_workbook(fake_xlsx)

    assert call_count == 2          # 1 failed + 1 succeeded
    assert result == {"id": "ok"}


@pytest.mark.asyncio
async def test_upload_protocol_workbook_raises_protokollo_locked_after_all_retries(
    client: OneDriveClient, tmp_path: Path, monkeypatch,
) -> None:
    """After exhausting retries, raises ProtokolloLockedError with a Greek message."""
    from src.integrations.m365.onedrive import (
        _PROTOKOLLO_LOCK_RETRY_DELAYS, ProtokolloLockedError,
    )

    async def _no_sleep(_): return None
    monkeypatch.setattr("asyncio.sleep", _no_sleep)

    async def fake_upload(**_):
        resp = httpx.Response(status_code=423, request=httpx.Request("PUT", "https://x"))
        raise httpx.HTTPStatusError("423", request=resp.request, response=resp)

    fake_xlsx = tmp_path / "p.xlsx"
    fake_xlsx.write_bytes(b"x")

    with patch.object(client, "upload_file", side_effect=fake_upload):
        with pytest.raises(ProtokolloLockedError) as excinfo:
            await client._upload_protocol_workbook(fake_xlsx)
    # Message should be the user-facing Greek string, not the raw HTTP error
    assert "κλειδωμένο" in str(excinfo.value)
    assert "Excel" in str(excinfo.value)


@pytest.mark.asyncio
async def test_upload_protocol_workbook_propagates_non_423_errors(
    client: OneDriveClient, tmp_path: Path,
) -> None:
    """500-class errors propagate immediately - no retry, no Protokollo wrapping."""
    async def fake_upload(**_):
        resp = httpx.Response(status_code=500, request=httpx.Request("PUT", "https://x"))
        raise httpx.HTTPStatusError("500", request=resp.request, response=resp)

    fake_xlsx = tmp_path / "p.xlsx"
    fake_xlsx.write_bytes(b"x")

    with patch.object(client, "upload_file", side_effect=fake_upload):
        with pytest.raises(httpx.HTTPStatusError) as excinfo:
            await client._upload_protocol_workbook(fake_xlsx)
    assert excinfo.value.response.status_code == 500
