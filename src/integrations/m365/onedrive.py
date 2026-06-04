"""Microsoft OneDrive / SharePoint integration via Microsoft Graph API.

Authentication model
--------------------
The bot uses **Delegated** permissions (not Application permissions).
This means:
  - The MSAL ``ConfidentialClientApplication`` is provided by the shared
    :class:`M365GraphAuthMixin` (see :mod:`src.integrations.m365.auth`).
  - The first sign-in is done interactively via the OAuth2 authorization-code
    flow (``authenticate_interactive``).  It opens the user's browser, listens
    on localhost:8000 for the redirect, and exchanges the code for tokens.
  - Subsequent calls use the MSAL token cache (refresh_token kept across
    restarts via ``src.core.tokens`` → ``data/tokens.json``).
  - If no valid cached token is found, ``M365AuthRequired`` (aliased as
    ``OneDriveAuthRequired`` for backwards compatibility) is raised.
    The caller should tell the user to run ``ai-assistant auth microsoft``.

Storage location
----------------
Files are archived in the **Board of Directors SharePoint site**:
  - Host:   amnestygr.sharepoint.com
  - Site:   /sites/Board
  - Drive:  the site's default document library (Shared Documents)
  - Root:   Αρχείο/

Graph URL pattern for file operations:
  GET/PUT /sites/{site-id}/drive/root:/{url-encoded-path}:/content
"""

from __future__ import annotations

import http.server
import logging
import threading
import urllib.parse
import webbrowser
from pathlib import Path
from typing import Any

import httpx

from src.config import settings
from src.core.audit import log_action
from src.integrations.m365.auth import M365AuthRequired, M365GraphAuthMixin

logger = logging.getLogger(__name__)

_GRAPH_BASE = "https://graph.microsoft.com/v1.0"


# Backwards-compatibility alias: callers that import ``OneDriveAuthRequired``
# (or ``raise``/``except`` it by name) keep working unchanged.  All new code
# should use :class:`M365AuthRequired` from :mod:`src.integrations.m365.auth`.
OneDriveAuthRequired = M365AuthRequired


class ProtokolloLockedError(RuntimeError):
    """Raised when the πρωτόκολλο xlsx is locked (someone has it open in Excel).

    SharePoint returns HTTP 423 Locked when another client holds an edit
    lock on a file.  We retry a few times with backoff in case it's a
    transient autosave lock, then surface this exception so callers can
    show a user-friendly Greek message instead of a raw HTTP error.
    """


# Per-attempt backoff schedule when retrying a 423 Locked write.  Three
# attempts total: immediate, +2s, +5s, +10s.  If still locked, give up.
# Chosen short enough to feel responsive to the user (~17s worst case)
# but long enough to ride out brief autosave windows.
_PROTOKOLLO_LOCK_RETRY_DELAYS = (2.0, 5.0, 10.0)


class OneDriveClient(M365GraphAuthMixin):
    """Client for SharePoint / OneDrive operations via Microsoft Graph API.

    The public API (upload_file, create_folder, list_files, download_file,
    get_share_link) targets the Amnesty Board SharePoint site and is backward-
    compatible with callers that pass paths relative to ``Αρχείο``.

    Inherits token cache + ``_get_token`` / ``_persist_cache`` from
    :class:`M365GraphAuthMixin`.  Overrides ``_headers`` to support a custom
    ``content_type`` (needed for binary uploads via ``application/octet-stream``).
    """

    # Delegated scopes needed for SharePoint / OneDrive file operations.
    # offline_access is requested implicitly by MSAL when the app is configured
    # with a token cache — do NOT include it in the scopes list passed to
    # acquire_token_silent / acquire_token_by_authorization_code (MSAL adds it).
    _SCOPES = ["Files.ReadWrite.All"]

    def __init__(self) -> None:
        super().__init__()
        # Site ID is resolved once per client instance and cached in memory.
        self._site_id: str | None = None

    # ── Headers override (binary uploads need application/octet-stream) ──────

    def _headers(self, content_type: str = "application/json") -> dict[str, str]:
        return {
            "Authorization": f"Bearer {self._get_token()}",
            "Content-Type": content_type,
        }

    # ── Interactive first-run sign-in ─────────────────────────────────────────

    def authenticate_interactive(self) -> None:
        """Interactive first-run sign-in: open browser, listen for redirect.

        Opens the Microsoft login page in the default browser, starts a
        one-shot HTTP server on localhost:8000 to receive the OAuth2 callback,
        exchanges the authorization code for tokens, and persists them.
        """
        redirect_uri = settings.ms_redirect_uri
        # Parse port from redirect URI (default 8000)
        parsed = urllib.parse.urlparse(redirect_uri)
        port = parsed.port or 8000

        auth_url = self._app.get_authorization_request_url(
            scopes=self._SCOPES,
            redirect_uri=redirect_uri,
        )

        # We use a mutable container to pass the code out of the handler.
        result_holder: dict[str, str] = {}
        server_ready = threading.Event()
        server_done = threading.Event()

        class _CallbackHandler(http.server.BaseHTTPRequestHandler):
            def do_GET(self) -> None:
                parsed_path = urllib.parse.urlparse(self.path)
                params = urllib.parse.parse_qs(parsed_path.query)
                code = params.get("code", [None])[0]
                error = params.get("error", [None])[0]

                if code:
                    result_holder["code"] = code
                    body = (
                        b"<html><body><h2>Authentication successful!</h2>"
                        b"<p>You can close this tab and return to the terminal.</p>"
                        b"</body></html>"
                    )
                else:
                    result_holder["error"] = error or "unknown"
                    body = (
                        b"<html><body><h2>Authentication failed.</h2>"
                        b"<p>Error: " + (error or "unknown").encode() + b"</p>"
                        b"</body></html>"
                    )

                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                # Signal the main thread to shut down the server.
                threading.Thread(target=server_done.set).start()

            def log_message(self, *args: Any) -> None:  # suppress request logs
                pass

        httpd = http.server.HTTPServer(("localhost", port), _CallbackHandler)

        def _serve() -> None:
            server_ready.set()
            # Serve until the callback handler sets server_done.
            while not server_done.is_set():
                httpd.handle_request()
            httpd.server_close()

        server_thread = threading.Thread(target=_serve, daemon=True)
        server_thread.start()
        server_ready.wait(timeout=5)

        logger.info("Opening browser for Microsoft sign-in: %s", auth_url)
        webbrowser.open(auth_url)

        # Wait for the callback (timeout: 5 minutes)
        if not server_done.wait(timeout=300):
            raise TimeoutError("Microsoft sign-in timed out (no callback received after 5 minutes).")

        if "error" in result_holder:
            raise RuntimeError(f"Microsoft sign-in error: {result_holder['error']}")

        code = result_holder["code"]
        token_result = self._app.acquire_token_by_authorization_code(
            code=code,
            scopes=self._SCOPES,
            redirect_uri=redirect_uri,
        )

        if "access_token" not in token_result:
            raise RuntimeError(
                f"Token exchange failed: {token_result.get('error_description', token_result)}"
            )

        self._persist_cache()
        logger.info("Microsoft sign-in successful — token cached.")

    # ── SharePoint site resolution ────────────────────────────────────────────

    def _graph_url(self, path: str) -> str:
        """Prepend the Graph base URL."""
        return f"{_GRAPH_BASE}{path}"

    def _encode_path(self, path: str) -> str:
        """URL-encode a SharePoint path, preserving forward slashes."""
        return urllib.parse.quote(path, safe="/")

    def _get_site_id(self) -> str:
        """Resolve and cache the SharePoint site ID.

        Uses GET /sites/{host}:{site_path} which returns a composite ID like
        "amnestygr.sharepoint.com,<site-guid>,<web-guid>".
        """
        if self._site_id:
            return self._site_id

        host = settings.onedrive.sharepoint_host
        site_path = settings.onedrive.sharepoint_site_path
        url = self._graph_url(f"/sites/{host}:{site_path}")

        with httpx.Client() as client:
            resp = client.get(url, headers=self._headers())
            resp.raise_for_status()

        self._site_id = resp.json()["id"]
        logger.debug("Resolved SharePoint site ID: %s", self._site_id)
        return self._site_id

    def _site_drive_url(self, relative_path: str) -> str:
        """Build a Graph URL for a path inside the site's default drive.

        Args:
            relative_path: Path relative to the drive root (e.g.
                "Αρχείο/Αρχείο ανά έτος/2026/file.pdf").
        """
        site_id = self._get_site_id()
        encoded = self._encode_path(relative_path)
        return self._graph_url(f"/sites/{site_id}/drive/root:/{encoded}")

    def _archive_path(self, *parts: str) -> str:
        """Join archive_root + caller-supplied parts into a single path."""
        components = [settings.onedrive.archive_root] + list(parts)
        # Strip leading/trailing slashes from each part before joining
        return "/".join(p.strip("/") for p in components if p.strip("/"))

    # ── Public API ────────────────────────────────────────────────────────────

    async def upload_file(
        self,
        local_path: Path,
        remote_folder: str,
        filename: str | None = None,
        workflow: str = "onedrive",
    ) -> dict[str, Any]:
        """Upload a file to the SharePoint archive.

        Args:
            local_path: Path to the local file.
            remote_folder: Folder path **relative to archive_root** (e.g.
                "Αρχείο ανά έτος/2026").
            filename: Override filename; defaults to local file name.
            workflow: Workflow name for audit logging.

        Returns:
            Graph API driveItem dict.
        """
        filename = filename or local_path.name
        remote_path = self._archive_path(remote_folder, filename)

        with open(local_path, "rb") as fh:
            content = fh.read()

        url = f"{self._site_drive_url(remote_path)}:/content"
        async with httpx.AsyncClient() as client:
            response = await client.put(
                url,
                headers=self._headers(content_type="application/octet-stream"),
                content=content,
            )
            response.raise_for_status()

        result = response.json()
        log_action(
            workflow=workflow,
            action="file_uploaded",
            actor="system",
            target=remote_path,
            details={"file_id": result.get("id"), "size": result.get("size")},
        )
        logger.info("Uploaded %s → SharePoint:%s", filename, remote_path)
        return result

    async def create_folder(self, folder_path: str) -> dict[str, Any]:
        """Create a folder in the SharePoint archive.

        Args:
            folder_path: Path **relative to archive_root**.

        Returns:
            Graph API driveItem dict, or ``{"status": "already_exists", ...}``.
        """
        remote_path = self._archive_path(folder_path)
        parent, name = remote_path.rsplit("/", 1)
        site_id = self._get_site_id()
        encoded_parent = self._encode_path(parent)
        url = self._graph_url(
            f"/sites/{site_id}/drive/root:/{encoded_parent}:/children"
        )

        async with httpx.AsyncClient() as client:
            response = await client.post(
                url,
                headers=self._headers(),
                json={
                    "name": name,
                    "folder": {},
                    "@microsoft.graph.conflictBehavior": "fail",
                },
            )
            if response.status_code == 409:
                return {"status": "already_exists", "path": remote_path}
            response.raise_for_status()
            return response.json()

    async def list_files(self, folder_path: str) -> list[dict[str, Any]]:
        """List items inside a SharePoint archive folder.

        Args:
            folder_path: Path **relative to archive_root** (use "" or "/" for
                the archive root itself).

        Returns:
            List of Graph API driveItem dicts.
        """
        # Allow callers to pass "" or "/" to mean "list the archive root"
        if folder_path.strip("/"):
            remote_path = self._archive_path(folder_path)
        else:
            remote_path = settings.onedrive.archive_root

        url = f"{self._site_drive_url(remote_path)}:/children"
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers=self._headers())
            response.raise_for_status()
            return response.json().get("value", [])

    async def download_file(self, remote_path: str, local_path: Path) -> Path:
        """Download a file from the SharePoint archive.

        Args:
            remote_path: File path **relative to archive_root**.
            local_path: Local destination path.

        Returns:
            Path to the saved local file.
        """
        full_path = self._archive_path(remote_path)
        url = f"{self._site_drive_url(full_path)}:/content"
        async with httpx.AsyncClient() as client:
            response = await client.get(
                url,
                headers=self._headers(),
                follow_redirects=True,
            )
            response.raise_for_status()
            local_path.write_bytes(response.content)
        logger.info("Downloaded SharePoint:%s → %s", full_path, local_path)
        return local_path

    async def get_share_link(self, file_id: str) -> str:
        """Create a sharing link for a file by its item ID.

        Args:
            file_id: Graph API driveItem ID.

        Returns:
            Web URL for the sharing link.
        """
        site_id = self._get_site_id()
        url = self._graph_url(
            f"/sites/{site_id}/drive/items/{file_id}/createLink"
        )
        async with httpx.AsyncClient() as client:
            response = await client.post(
                url,
                headers=self._headers(),
                json={"type": "view", "scope": "organization"},
            )
            response.raise_for_status()
            return response.json()["link"]["webUrl"]

    async def delete_file(self, remote_path: str, workflow: str = "onedrive") -> None:
        """Delete a file from the SharePoint archive.

        Args:
            remote_path: Path **relative to archive_root** (e.g.
                ``"Αρχείο ανά έτος/2026/[2026_017] Πρόσκληση.pdf"``).
            workflow:    Workflow name for audit logging.

        Raises:
            httpx.HTTPStatusError if the item does not exist (404) or the
            caller does not have permission (403).
        """
        full_path = self._archive_path(remote_path)
        url = f"{self._site_drive_url(full_path)}"
        async with httpx.AsyncClient() as client:
            response = await client.delete(url, headers=self._headers())
            response.raise_for_status()
        log_action(
            workflow=workflow,
            action="file_deleted",
            actor="system",
            target=full_path,
            details={"status": response.status_code},
        )
        logger.info("Deleted SharePoint:%s", full_path)

    async def delete_folder(self, folder_path: str, workflow: str = "onedrive") -> None:
        """Delete a folder (recursively) from the SharePoint archive.

        Same Graph endpoint as :meth:`delete_file` — Graph treats folder
        deletes recursively by default.  Kept as a separate method for call-
        site clarity; use whichever reads better at the call site.

        Args:
            folder_path: Path **relative to archive_root**.
            workflow:    Workflow name for audit logging.
        """
        full_path = self._archive_path(folder_path)
        url = f"{self._site_drive_url(full_path)}"
        async with httpx.AsyncClient() as client:
            response = await client.delete(url, headers=self._headers())
            response.raise_for_status()
        log_action(
            workflow=workflow,
            action="folder_deleted",
            actor="system",
            target=full_path,
            details={"status": response.status_code},
        )
        logger.info("Deleted SharePoint folder:%s", full_path)

    # ── Protocol registry (Excel) ─────────────────────────────────────────────

    async def get_next_protocol_number(self, year: int) -> str:
        """Return the next protocol number for *year* by reading the Excel registry.

        Downloads ``config.onedrive.protocol_excel`` from the archive root,
        inspects the year's sheet, finds the last ``YYYY_NNN`` entry, and
        increments the sequence.

        Returns:
            ``"YYYY_NNN"`` string, e.g. ``"2026_026"``.
            Falls back to ``"YYYY_001"`` if the sheet is empty or missing.
        """
        import re as _re

        import openpyxl

        wb_path = await self._workbook_path_for_read()
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)

        sheet_name = str(year)
        if sheet_name not in wb.sheetnames:
            return f"{year}_001"

        ws = wb[sheet_name]
        last_protocol: str | None = None
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val and _re.match(r"^\d{4}[-_]\d+$", str(val).strip()):
                last_protocol = str(val).strip()

        if not last_protocol:
            return f"{year}_001"

        m = _re.match(r"^(\d{4})[-_](\d+)$", last_protocol)
        if m:
            entry_year = int(m.group(1))
            seq = int(m.group(2))
            if entry_year < year:
                return f"{year}_001"
            return f"{year}_{seq + 1:03d}"

        return f"{year}_001"

    async def append_protocol_row(
        self,
        protocol_id: str,
        date_str: str,
        title: str,
        main_points: str = "",
        tags: str = "",
    ) -> None:
        """Append a new row to the protocol Excel registry and re-upload.

        Downloads ``config.onedrive.protocol_excel``, adds the row to the
        appropriate year sheet (creating it with a header if absent), and
        uploads the modified workbook back to the archive root.

        Args:
            protocol_id: e.g. ``"2026_026"``
            date_str:    ISO date string ``"YYYY-MM-DD"`` (or ``""``).
            title:       Document title (e.g. ``"Πρόσκληση - Συνεδρίαση ΔΣ05-2026"``).
            main_points: Optional bullet points / summary (can be multi-line).
            tags:        Comma-separated Ετικέτες string.
        """
        import re as _re
        import tempfile
        from datetime import date as _date
        from datetime import datetime as _dt

        import openpyxl

        year_match = _re.match(r"^(\d{4})", protocol_id)
        year = int(year_match.group(1)) if year_match else _date.today().year

        excel_filename = settings.onedrive.protocol_excel
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = Path(f.name)
        try:
            await self.download_file(excel_filename, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)

            sheet_name = str(year)
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(["Πρωτόκολλο", "Ημερομηνία", "Έγγραφο", "Κύρια Σημεία", "Ετικέτες"])
            else:
                ws = wb[sheet_name]

            # Parse date string → Python date object so Excel stores it as a date cell
            date_val: _date | None = None
            if date_str:
                try:
                    date_val = _dt.strptime(date_str[:10], "%Y-%m-%d").date()
                except ValueError:
                    pass

            ws.append([protocol_id, date_val, title, main_points or None, tags or None])
            wb.save(tmp_path)

            # Re-upload to the archive root with retry-on-423 (handles the
            # case where someone has the xlsx open in Excel).
            await self._upload_protocol_workbook(tmp_path, workflow="onedrive")
            logger.info("Protocol registry updated: %s — %s", protocol_id, title)
        finally:
            try:
                tmp_path.unlink()
            except OSError:
                pass

    async def update_protocol_row(
        self,
        protocol_id: str,
        *,
        title: str | None = None,
        main_points: str | None = None,
        tags: str | None = None,
    ) -> bool:
        """Modify columns of an existing πρωτόκολλο row in-place.

        Each kwarg that is *not* ``None`` overwrites the corresponding column;
        passing ``None`` keeps the existing value.  Pass ``""`` to clear a
        cell explicitly.  Date and the πρωτόκολλο id itself are never touched
        by this method — those are write-once.

        Args:
            protocol_id:  Row key, e.g. ``"2026_017"``.
            title:        New title (column C) or None to leave alone.
            main_points:  New Κύρια Σημεία (column D).
            tags:         New Ετικέτες comma-string (column E).

        Returns:
            ``True`` if the row was found and updated; ``False`` if no
            matching row exists (no-op).
        """
        import re as _re
        import tempfile
        import openpyxl

        if title is None and main_points is None and tags is None:
            # No-op — don't bother downloading.
            return False

        year_match = _re.match(r"^(\d{4})", protocol_id)
        if not year_match:
            return False
        year = year_match.group(1)

        excel_filename = settings.onedrive.protocol_excel
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = Path(f.name)
        try:
            await self.download_file(excel_filename, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            if year not in wb.sheetnames:
                return False
            ws = wb[year]

            # Find the row whose column A matches protocol_id.
            target_row: int | None = None
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2
            ):
                if row[0] and str(row[0]).strip() == protocol_id.strip():
                    target_row = row_idx
                    break
            if target_row is None:
                return False

            # Columns: A=Πρωτόκολλο, B=Ημερομηνία, C=Έγγραφο, D=Κύρια Σημεία, E=Ετικέτες
            if title is not None:
                ws.cell(row=target_row, column=3).value = title
            if main_points is not None:
                ws.cell(row=target_row, column=4).value = main_points or None
            if tags is not None:
                ws.cell(row=target_row, column=5).value = tags or None

            wb.save(tmp_path)

            await self._upload_protocol_workbook(tmp_path, workflow="onedrive")
            logger.info(
                "Protocol row updated: %s (row %d) fields=%s",
                protocol_id, target_row,
                [k for k, v in (("title", title), ("main_points", main_points),
                                ("tags", tags)) if v is not None],
            )
            return True
        finally:
            try:
                tmp_path.unlink()
            except OSError:
                pass

    async def rename_file(
        self,
        remote_path: str,
        new_name: str,
        *,
        workflow: str = "onedrive",
    ) -> dict[str, Any]:
        """Rename a SharePoint file in place (same folder, new name).

        Uses Graph's ``PATCH /drive/items/...`` with a ``{"name": ...}`` body.
        The file's Graph ``id`` is preserved — share links etc. keep working.

        Args:
            remote_path: Current path **relative to archive_root**
                (e.g. ``"Αρχείο ανά έτος/2026/[2026_017] Old.pdf"``).
            new_name:    New filename (just the leaf — no directory).  Caller
                is responsible for sanitising filesystem-hostile characters.
            workflow:    Audit-log workflow name.

        Returns:
            The Graph ``driveItem`` dict for the renamed file.
        """
        full_path = self._archive_path(remote_path)
        url = self._site_drive_url(full_path)
        async with httpx.AsyncClient() as client:
            response = await client.patch(
                url,
                headers=self._headers(),
                json={"name": new_name},
            )
            response.raise_for_status()
            data = response.json()
        log_action(
            workflow=workflow,
            action="file_renamed",
            actor="system",
            target=full_path,
            details={"new_name": new_name, "id": data.get("id")},
        )
        logger.info("Renamed SharePoint:%s → %s", full_path, new_name)
        return data

    async def file_exists_for_protocol(self, protocol_id: str) -> bool:
        """True if ANY file in ``Αρχείο ανά έτος/{year}/`` starts with ``[{protocol_id}] ``.

        Used by the archive workflow's pre-existence check to decide whether
        a πρωτόκολλο row is "reserved-but-empty" (SecGen pre-claimed the
        number, no file yet) or "already archived" (row + file both exist —
        the bot must refuse to overwrite).

        Args:
            protocol_id: e.g. ``"2026_017"``.

        Returns:
            True if at least one matching file is present, False otherwise.
            Network errors return False (fail-open — we'd rather attempt a
            write and have SharePoint reject than refuse on a transient
            list failure).
        """
        import re as _re
        year_match = _re.match(r"^(\d{4})", protocol_id)
        if not year_match:
            return False
        year = year_match.group(1)
        folder = f"{settings.onedrive.yearly_subfolder}/{year}"
        prefix = f"[{protocol_id}] "
        try:
            items = await self.list_files(folder)
        except Exception as exc:
            logger.warning(
                "file_exists_for_protocol: list_files(%s) failed (non-fatal, "
                "returning False): %s",
                folder, exc,
            )
            return False
        for item in items:
            name = (item.get("name") or "")
            if name.startswith(prefix):
                return True
        return False

    async def find_protocol_row(self, protocol_id: str) -> dict[str, str] | None:
        """Look up a row in the πρωτόκολλο xlsx by protocol_id.

        Returns:
            Dict with keys ``proto``, ``date``, ``title``, ``key_points``,
            ``tags`` if found, else None.  Used by the Phase 4 collision
            gate (and any future read-only queries that need a single row).
        """
        import re as _re
        import openpyxl

        year_match = _re.match(r"^(\d{4})", protocol_id)
        if not year_match:
            return None
        year = year_match.group(1)

        wb_path = await self._workbook_path_for_read()
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        if year not in wb.sheetnames:
            return None
        ws = wb[year]
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            cell_proto = row[0]
            if cell_proto and str(cell_proto).strip() == protocol_id.strip():
                date_val = row[1] if len(row) > 1 else None
                return {
                    "proto": str(cell_proto).strip(),
                    "date": (
                        date_val.isoformat()
                        if hasattr(date_val, "isoformat")
                        else (str(date_val).strip() if date_val is not None else "")
                    ),
                    "title": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
                    "key_points": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                    "tags": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
                }
        return None

    async def delete_protocol_row(self, protocol_id: str) -> bool:
        """Remove a row from the protocol Excel registry by its protocol_id.

        Used by ``ai-assistant invite --cancel`` to wipe an invitation entry
        when a meeting is cancelled.  Year is parsed from the protocol_id
        prefix (``"2026_017"`` → sheet ``"2026"``).

        Args:
            protocol_id: e.g. ``"2026_017"``.

        Returns:
            ``True`` if a matching row was found and removed; ``False`` if
            the year sheet or the row is not present (no-op, not an error).
        """
        import re as _re
        import tempfile

        import openpyxl

        year_match = _re.match(r"^(\d{4})", protocol_id)
        if not year_match:
            return False
        year = year_match.group(1)

        excel_filename = settings.onedrive.protocol_excel
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = Path(f.name)
        try:
            await self.download_file(excel_filename, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            if year not in wb.sheetnames:
                return False
            ws = wb[year]

            # Find the first row whose column A matches protocol_id.
            target_row: int | None = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
                if row[0] and str(row[0]).strip() == protocol_id.strip():
                    target_row = row_idx
                    break
            if target_row is None:
                return False

            ws.delete_rows(target_row, 1)
            wb.save(tmp_path)

            await self._upload_protocol_workbook(tmp_path, workflow="onedrive")
            logger.info("Protocol registry row removed: %s (year %s, row %d)", protocol_id, year, target_row)
            return True
        finally:
            try:
                tmp_path.unlink()
            except OSError:
                pass

    # ── Taxonomy / categories reference data (Phase 1 archive) ────────────────

    # Local safety copy of the πρωτόκολλο xlsx — refreshed on every successful
    # download.  Defends against accidental deletion / corruption of the file in
    # OneDrive: the most recent known-good copy is always one ``cp`` away.
    PROTOCOL_BACKUP_PATH: Path = Path("data") / "backups" / "protokollo_latest.xlsx"

    def _backup_protocol_workbook(self, src: Path) -> None:
        """Atomically replace the local backup copy of the πρωτόκολλο xlsx.

        Called from :meth:`_download_protocol_workbook` after each successful
        download.  Best-effort — failures here NEVER propagate (the running
        workflow already has its tmp copy and shouldn't crash because of a
        backup hiccup).  The atomic replace pattern (copy → temp sibling →
        ``os.replace``) means we never leave a half-written backup behind.
        """
        try:
            dest = self.PROTOCOL_BACKUP_PATH
            dest.parent.mkdir(parents=True, exist_ok=True)
            # Write to a sibling temp then atomically rename — guarantees the
            # backup is either the previous version or the new one, never
            # partial.
            tmp_sibling = dest.with_suffix(dest.suffix + ".tmp")
            import shutil
            shutil.copy2(src, tmp_sibling)
            import os
            os.replace(tmp_sibling, dest)
            logger.debug("Refreshed πρωτόκολλο backup at %s (%d bytes)",
                         dest, dest.stat().st_size)
        except Exception as e:  # pragma: no cover — best-effort
            logger.warning("Failed to refresh πρωτόκολλο backup: %s", e)

    async def _download_protocol_workbook(self) -> Path:
        """Download the πρωτόκολλο xlsx to a tempfile and return its path.

        Side-effect: refreshes the local safety backup at
        :attr:`PROTOCOL_BACKUP_PATH` on every successful download.

        Caller is responsible for deleting the tempfile.

        Used by **write** paths (append/update/delete row) which must always
        operate on the latest server state and then PUT back.  Read paths
        should go through :meth:`_workbook_path_for_read` so that within a
        single workflow run all reads share one snapshot.
        """
        import tempfile

        excel_filename = settings.onedrive.protocol_excel
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = Path(f.name)
        await self.download_file(excel_filename, tmp_path)
        # Now that the download succeeded, refresh the local safety copy.
        self._backup_protocol_workbook(tmp_path)
        return tmp_path

    async def refresh_protocol_workbook(self) -> Path:
        """Download the πρωτόκολλο xlsx and update the local snapshot.

        Call this **once** at the start of a workflow run to take a fresh
        snapshot.  All subsequent reads (taxonomy, categories, recent
        entries, row lookups, max-seq) will reuse that snapshot through
        :meth:`_workbook_path_for_read` — one network download per run
        instead of one per read.

        Returns the path to the refreshed backup (so callers don't need to
        reach for :attr:`PROTOCOL_BACKUP_PATH` directly).  The caller MUST
        NOT delete this file — it's the shared snapshot.
        """
        tmp_path = await self._download_protocol_workbook()
        # tmp_path was copied into PROTOCOL_BACKUP_PATH as a side effect of
        # _download_protocol_workbook(); the tmp itself can go.
        try:
            tmp_path.unlink()
        except OSError:
            pass
        return self.PROTOCOL_BACKUP_PATH

    async def _workbook_path_for_read(self) -> Path:
        """Return a path to the πρωτόκολλο xlsx for **read-only** use.

        Reuses :attr:`PROTOCOL_BACKUP_PATH` if it exists (refreshed by the
        most recent :meth:`refresh_protocol_workbook` / write call).  Falls
        back to a fresh download only when no backup is present yet (first
        run on a new machine).

        Callers MUST NOT delete the returned path — it's the shared
        snapshot for the current workflow run.
        """
        if self.PROTOCOL_BACKUP_PATH.exists():
            return self.PROTOCOL_BACKUP_PATH
        # First run: no backup yet, download once (side-effect creates backup).
        logger.info(
            "πρωτόκολλο backup missing — taking initial snapshot at %s",
            self.PROTOCOL_BACKUP_PATH,
        )
        return await self.refresh_protocol_workbook()

    async def _upload_protocol_workbook(
        self,
        local_path: Path,
        *,
        workflow: str = "onedrive",
    ) -> dict[str, Any]:
        """PUT the πρωτόκολλο xlsx back to SharePoint with retry-on-423.

        SharePoint returns 423 Locked when someone has the file open in
        Excel desktop or Excel Online with an edit lock.  This usually
        clears within a few seconds (autosave window), but can persist
        as long as the user has the file open.

        Strategy: try the PUT; on 423, wait + retry up to len(
        _PROTOKOLLO_LOCK_RETRY_DELAYS) extra times; if still locked, raise
        :class:`ProtokolloLockedError` so the caller can surface a clean
        Greek message to the user instead of the raw HTTP exception.

        Other HTTP errors propagate unchanged.
        """
        import asyncio as _asyncio
        import httpx as _httpx

        excel_filename = settings.onedrive.protocol_excel
        attempts = [0.0, *_PROTOKOLLO_LOCK_RETRY_DELAYS]
        last_exc: Exception | None = None
        for delay in attempts:
            if delay > 0:
                logger.info(
                    "πρωτόκολλο xlsx locked — retrying in %.1fs (attempt %d/%d)",
                    delay, attempts.index(delay), len(attempts) - 1,
                )
                await _asyncio.sleep(delay)
            try:
                return await self.upload_file(
                    local_path=local_path,
                    remote_folder="",
                    filename=excel_filename,
                    workflow=workflow,
                )
            except _httpx.HTTPStatusError as exc:
                if exc.response.status_code != 423:
                    raise  # Non-lock error → propagate immediately
                last_exc = exc
                continue
        # All attempts exhausted — give up with a clean error
        raise ProtokolloLockedError(
            "Το πρωτόκολλο xlsx είναι κλειδωμένο (πιθανώς το έχει ανοιχτό κάποιος "
            "στο Excel). Παρακαλώ κλείστε το από Excel (επιφάνεια εργασίας ή web) "
            "και ξαναπροσπαθήστε."
        ) from last_exc

    @staticmethod
    def _looks_like_header(value: Any) -> bool:
        """Heuristic for skipping a header row.

        Header cells in the user's xlsx are Greek labels like 'Ετικέτα',
        'Περιγραφή', 'Πρότυπο τίτλου', 'Προεπιλεγμένες Ετικέτες', etc.
        They never look like a tag value (single Greek word) or a pattern
        (typically contains a Greek noun or '-').  Cheap, safe rule:
        skip the first row if its first cell text-matches the known header
        labels (case-insensitively, stripped).
        """
        if value is None:
            return False
        text = str(value).strip().lower()
        if not text:
            return False
        # Known Greek header keywords across both tabs.
        return any(
            kw in text
            for kw in (
                "ετικέτ", "περιγραφ", "πρότυπ", "τίτλ", "κατηγορ",
                "tag", "label", "description", "pattern", "title",
                "σημεί",
            )
        )

    # ── Workbook parsers (separate from download) ────────────────────────────
    #
    # These take an already-loaded openpyxl workbook so callers can amortise a
    # single network download over multiple sheet reads.  The async public
    # methods below are thin convenience wrappers around them.

    def _parse_taxonomy(self, wb: Any) -> list[dict[str, str]]:
        """Pull rows from the ``Ετικέτες`` tab of an open workbook."""
        if "Ετικέτες" not in wb.sheetnames:
            return []
        ws = wb["Ετικέτες"]
        rows = list(ws.iter_rows(min_row=1, max_col=2, values_only=True))
        if not rows:
            return []
        if self._looks_like_header(rows[0][0]):
            rows = rows[1:]
        out: list[dict[str, str]] = []
        for row in rows:
            tag = (str(row[0]).strip() if row[0] is not None else "")
            desc = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
            if tag:
                out.append({"tag": tag, "description": desc})
        return out

    def _parse_categories(self, wb: Any) -> list[dict[str, str]]:
        """Pull rows from the ``Κατηγορίες`` tab of an open workbook."""
        if "Κατηγορίες" not in wb.sheetnames:
            return []
        ws = wb["Κατηγορίες"]
        rows = list(ws.iter_rows(min_row=1, max_col=3, values_only=True))
        if not rows:
            return []
        if self._looks_like_header(rows[0][0]):
            rows = rows[1:]
        out: list[dict[str, str]] = []
        for row in rows:
            pattern = (str(row[0]).strip() if row[0] is not None else "")
            tags = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
            kuria = (str(row[2]).strip() if len(row) > 2 and row[2] is not None else "")
            if pattern:
                out.append({"pattern": pattern, "tags": tags, "kuria_simeia": kuria})
        return out

    async def read_taxonomy_and_categories(
        self,
    ) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
        """Read both ``Ετικέτες`` and ``Κατηγορίες`` in ONE network round-trip.

        Used by the archive workflow's classification step to avoid downloading
        the (~100 KB) xlsx twice in a row.  Equivalent to::

            tax = await client.read_taxonomy()
            cat = await client.read_categories()

        but with half the HTTP traffic and roughly half the latency.
        """
        import openpyxl

        wb_path = await self._workbook_path_for_read()
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        return self._parse_taxonomy(wb), self._parse_categories(wb)

    async def read_taxonomy(self) -> list[dict[str, str]]:
        """Read the ``Ετικέτες`` tab and return ``[{tag, description}, ...]``.

        Convenience wrapper that downloads the xlsx and parses only the
        Ετικέτες tab.  When you need BOTH tabs in the same call, prefer
        :meth:`read_taxonomy_and_categories` to avoid a duplicate download.
        """
        import openpyxl

        wb_path = await self._workbook_path_for_read()
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        return self._parse_taxonomy(wb)

    async def read_categories(self) -> list[dict[str, str]]:
        """Read the ``Κατηγορίες`` tab and return canonical patterns.

        Convenience wrapper.  See :meth:`read_taxonomy_and_categories` to
        share a download with the taxonomy read.
        """
        import openpyxl

        wb_path = await self._workbook_path_for_read()
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        return self._parse_categories(wb)

    async def read_recent_entries(
        self,
        n: int = 30,
        years_back: int = 2,
    ) -> list[dict[str, str]]:
        """Read the most-recent *n* protocol entries across the last few years.

        Scans the current year's tab plus *years_back* prior years (so e.g.
        ``years_back=2`` reads {current, current-1}).  Picks the LAST *n*
        entries in chronological order (most recent first in the returned list).

        Args:
            n:           Maximum number of entries to return.
            years_back:  How many prior years to include (default 2 → current + 1 prior).

        Returns:
            List of dicts with keys ``proto``, ``date``, ``title``,
            ``key_points``, ``tags``.  Empty list if nothing found.
        """
        import openpyxl
        from datetime import date as _date

        current_year = _date.today().year
        # Build [current, current-1, ..., current-(years_back-1)]
        years = [str(current_year - i) for i in range(years_back)]

        wb_path = await self._workbook_path_for_read()
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        collected: list[dict[str, str]] = []
        # Iterate oldest year first so newer entries end up at the end of
        # the accumulator, then take the last n.
        for year in reversed(years):
            if year not in wb.sheetnames:
                continue
            ws = wb[year]
            # Header row is row 1; data starts at row 2.
            for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
                proto = row[0]
                if not proto:
                    continue
                date_val = row[1] if len(row) > 1 else None
                title = row[2] if len(row) > 2 else None
                key_points = row[3] if len(row) > 3 else None
                tags = row[4] if len(row) > 4 else None
                collected.append({
                    "proto": str(proto).strip(),
                    "date": (
                        date_val.isoformat()
                        if hasattr(date_val, "isoformat")
                        else (str(date_val).strip() if date_val is not None else "")
                    ),
                    "title": str(title).strip() if title is not None else "",
                    "key_points": str(key_points).strip() if key_points is not None else "",
                    "tags": str(tags).strip() if tags is not None else "",
                })
        # Last n entries (most recent end of the list)
        return collected[-n:] if collected else []

    async def get_current_year_max_seq(self, year: int) -> int:
        """Return the highest seq number present in the xlsx for *year*, or 0.

        Lighter than ``get_next_protocol_number`` for callers that want to
        seed ``reserve_next_protocol_number`` with the xlsx state.
        """
        import re as _re
        import openpyxl

        wb_path = await self._workbook_path_for_read()
        wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
        sheet_name = str(year)
        if sheet_name not in wb.sheetnames:
            return 0
        ws = wb[sheet_name]
        max_seq = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val:
                m = _re.match(r"^(\d{4})[-_](\d+)$", str(val).strip())
                if m and int(m.group(1)) == year:
                    max_seq = max(max_seq, int(m.group(2)))
        return max_seq
